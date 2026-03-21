import argparse
import hashlib
import json
import os
import sys
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal

if __package__ is None or __package__ == "":
    sys.path.append(str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from psycopg2.extras import Json

from src.db import db_cursor

CUSTOMER_SHEETS = (
    "customers_all",
    "leads_active",
    "leads_archived",
    "clients_active",
    "clients_archived",
)

COMMUNICATION_SHEET = "communications"
DEFAULT_SCHEMA_SQL = "/home/kv145/traffic-analytics/sql/045_alfacrm_crm_ingest.sql"


def norm_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def to_int(value):
    if value is None:
        return None
    try:
        text = str(value).strip()
        if text == "":
            return None
        return int(float(text))
    except Exception:
        return None


def to_smallint(value):
    i = to_int(value)
    if i is None:
        return None
    return i


def row_fingerprint(row):
    raw = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def file_md5(path):
    h = hashlib.md5()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def iter_sheet_dict_rows(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return
    ws = workbook[sheet_name]
    rows = ws.iter_rows(values_only=True)
    first = next(rows, None)
    if not first:
        return

    headers = [str(x).strip() if x is not None else "" for x in first]
    if len(headers) == 1 and headers[0].lower() == "empty":
        return

    for values in rows:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            val = values[idx] if idx < len(values) else None
            row[header] = norm_cell(val)
        if row:
            yield row


def apply_schema(cur, schema_sql_path):
    with open(schema_sql_path, "r", encoding="utf-8") as f:
        cur.execute(f.read())


def upsert_customer_row(cur, report_date, segment, source_file, row):
    customer_id = to_int(row.get("id"))
    if customer_id is None:
        return False

    customer_name = row.get("name")
    phone_normalized = row.get("phone_normalized") or row.get("phone")
    email_normalized = row.get("email_normalized") or row.get("email")
    telegram_username = row.get("telegram_username")
    is_study = to_smallint(row.get("is_study"))
    removed = to_smallint(row.get("removed"))

    cur.execute(
        """
        insert into stg_alfacrm_customers_daily (
            report_date,
            segment,
            customer_id,
            customer_name,
            phone_normalized,
            email_normalized,
            telegram_username,
            is_study,
            removed,
            source_file,
            payload_json,
            loaded_at
        )
        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now())
        on conflict (report_date, segment, customer_id)
        do update set
            customer_name = excluded.customer_name,
            phone_normalized = excluded.phone_normalized,
            email_normalized = excluded.email_normalized,
            telegram_username = excluded.telegram_username,
            is_study = excluded.is_study,
            removed = excluded.removed,
            source_file = excluded.source_file,
            payload_json = excluded.payload_json,
            loaded_at = now()
        """,
        (
            report_date,
            segment,
            customer_id,
            customer_name,
            phone_normalized,
            email_normalized,
            telegram_username,
            is_study,
            removed,
            source_file,
            Json(row),
        ),
    )
    return True


def upsert_communication_row(cur, report_date, source_file, row):
    communication_id = to_int(row.get("id") or row.get("communication_id"))
    customer_id = to_int(row.get("customer_id") or row.get("related_id"))
    communication_type = row.get("type") or row.get("communication_type") or row.get("class")
    created_at = row.get("created_at") or row.get("date") or row.get("add_date")

    if communication_id is not None:
        row_key = f"id:{communication_id}"
    else:
        row_key = f"fp:{row_fingerprint(row)}"

    cur.execute(
        """
        insert into stg_alfacrm_communications_daily (
            report_date,
            row_key,
            communication_id,
            customer_id,
            communication_type,
            created_at,
            source_file,
            payload_json,
            loaded_at
        )
        values (%s, %s, %s, %s, %s, %s, %s, %s, now())
        on conflict (report_date, row_key)
        do update set
            communication_id = excluded.communication_id,
            customer_id = excluded.customer_id,
            communication_type = excluded.communication_type,
            created_at = excluded.created_at,
            source_file = excluded.source_file,
            payload_json = excluded.payload_json,
            loaded_at = now()
        """,
        (
            report_date,
            row_key,
            communication_id,
            customer_id,
            communication_type,
            str(created_at) if created_at is not None else None,
            source_file,
            Json(row),
        ),
    )
    return True


def run(xlsx_path, report_date, source_file=None, skip_communications=False, schema_sql_path=None):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    source_file = source_file or os.path.basename(xlsx_path)
    file_hash = file_md5(xlsx_path)

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)

    customers_rows = 0
    communications_rows = 0

    with db_cursor() as (_, cur):
        if schema_sql_path:
            apply_schema(cur, schema_sql_path)

        for segment in CUSTOMER_SHEETS:
            for row in iter_sheet_dict_rows(wb, segment) or []:
                if upsert_customer_row(cur, report_date, segment, source_file, row):
                    customers_rows += 1

        if not skip_communications:
            for row in iter_sheet_dict_rows(wb, COMMUNICATION_SHEET) or []:
                if upsert_communication_row(cur, report_date, source_file, row):
                    communications_rows += 1

        cur.execute(
            """
            insert into etl_alfacrm_file_loads (
                report_date,
                source_file,
                file_hash,
                customers_rows,
                communications_rows,
                note
            )
            values (%s, %s, %s, %s, %s, %s)
            on conflict (file_hash)
            do update set
                report_date = excluded.report_date,
                source_file = excluded.source_file,
                customers_rows = greatest(etl_alfacrm_file_loads.customers_rows, excluded.customers_rows),
                communications_rows = greatest(etl_alfacrm_file_loads.communications_rows, excluded.communications_rows),
                note = excluded.note,
                loaded_at = now()
            """,
            (
                report_date,
                source_file,
                file_hash,
                customers_rows,
                communications_rows,
                "load_alfacrm_crm_xlsx",
            ),
        )

    wb.close()

    return {
        "status": "ok",
        "report_date": report_date,
        "source_file": source_file,
        "file_hash": file_hash,
        "customers_rows": customers_rows,
        "communications_rows": communications_rows,
    }


def parse_args():
    parser = argparse.ArgumentParser(description="Load AlfaCRM exporter XLSX into traffic-analytics staging")
    parser.add_argument("--xlsx", required=True, help="Path to alfacrm exporter xlsx file")
    parser.add_argument("--report-date", default=date.today().isoformat(), help="Logical report date YYYY-MM-DD")
    parser.add_argument("--source-file", default="", help="Optional source file label")
    parser.add_argument("--skip-communications", action="store_true", help="Skip communications sheet")
    parser.add_argument(
        "--schema-sql",
        default=DEFAULT_SCHEMA_SQL,
        help="SQL schema file to apply before load; pass empty string to skip",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    result = run(
        xlsx_path=args.xlsx,
        report_date=args.report_date,
        source_file=args.source_file or None,
        skip_communications=args.skip_communications,
        schema_sql_path=args.schema_sql or None,
    )
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
